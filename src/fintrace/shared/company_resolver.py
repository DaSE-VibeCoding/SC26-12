"""Resolve an exact stock code against the versioned local company master."""

from __future__ import annotations

import re
from collections.abc import Iterable
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from fintrace.shared.exceptions import (
    CompanyMasterDataUnavailableError,
    CompanyNotFoundError,
    InvalidCompanyCodeError,
)
from fintrace.shared.models import Company
from fintrace.shared.paths import require_file

COMPANY_CODE_PATTERN = re.compile(r"^(?P<code>\d{6})(?:\.(?P<suffix>SH|SZ|BJ))?$", re.IGNORECASE)
SUFFIX_EXCHANGE = {"SH": "SSE", "SZ": "SZSE", "BJ": "BSE"}


def infer_exchange(company_code: str) -> str:
    if company_code.startswith(("4", "8", "92")):
        return "BSE"
    if company_code.startswith(("5", "6", "9")):
        return "SSE"
    if company_code.startswith(("0", "1", "2", "3")):
        return "SZSE"
    raise InvalidCompanyCodeError(
        f"Cannot infer an exchange from stock code: {company_code}",
        step="normalize_company_code",
    )


def normalize_company_code(raw_code: str) -> tuple[str, str]:
    value = str(raw_code).strip().upper()
    match = COMPANY_CODE_PATTERN.fullmatch(value)
    if not match:
        raise InvalidCompanyCodeError(
            "Company code must be six digits with an optional .SH, .SZ, or .BJ suffix.",
            step="normalize_company_code",
            details={"input": value},
        )
    company_code = match.group("code")
    inferred_exchange = infer_exchange(company_code)
    suffix = match.group("suffix")
    if suffix and SUFFIX_EXCHANGE[suffix] != inferred_exchange:
        raise InvalidCompanyCodeError(
            f"Suffix .{suffix} conflicts with stock code {company_code}.",
            step="normalize_company_code",
            details={"inferred_exchange": inferred_exchange},
        )
    return company_code, inferred_exchange


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _symbol(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return f"{int(value):06d}"
    return str(value).strip().split(".")[0].zfill(6)


@lru_cache(maxsize=8)
def _rows_for_company(master_path: str, company_code: str) -> tuple[dict[str, Any], ...]:
    path = Path(master_path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        # Some exported workbooks declare only column A despite containing a full table.
        if sheet.max_column == 1:
            sheet.reset_dimensions()
        rows = sheet.iter_rows(values_only=True)
        headers = tuple(str(value) for value in next(rows))
        matched = tuple(
            dict(zip(headers, row, strict=False)) for row in rows if _symbol(row[0]) == company_code
        )
        workbook.close()
        return matched
    except (OSError, ValueError, StopIteration) as exc:
        raise CompanyMasterDataUnavailableError(
            f"Could not read company master data: {path}",
            step="load_company_master",
            details={"reason": str(exc)},
        ) from exc


def _select_version(rows: Iterable[dict[str, Any]], target_year: int) -> tuple[dict[str, Any], int]:
    dated_rows = [(row, _date(row.get("EndDate"))) for row in rows]
    candidates = [(row, end_date) for row, end_date in dated_rows if end_date is not None]
    same_year = [(row, end_date) for row, end_date in candidates if end_date.year == target_year]
    if same_year:
        selected, selected_date = max(same_year, key=lambda item: item[1])
        return selected, selected_date.year
    prior = [(row, end_date) for row, end_date in candidates if end_date.year < target_year]
    if prior:
        selected, selected_date = max(prior, key=lambda item: item[1])
        return selected, selected_date.year
    raise CompanyMasterDataUnavailableError(
        f"No company master version is available for or before {target_year}.",
        step="select_company_version",
    )


def resolve_company(raw_code: str, target_year: int) -> Company:
    if target_year < 1990 or target_year > 2100:
        raise InvalidCompanyCodeError(
            f"Target year is outside the supported range: {target_year}",
            step="validate_target_year",
        )
    company_code, exchange = normalize_company_code(raw_code)
    master_path = require_file("company_master_file")
    rows = _rows_for_company(str(master_path), company_code)
    if not rows:
        raise CompanyNotFoundError(
            f"Company code does not exist in local master data: {company_code}",
            step="match_company_code",
            details={"company_code": company_code},
        )
    row, used_year = _select_version(rows, target_year)
    fallback = used_year != target_year
    fallback_reason = None
    if fallback:
        fallback_reason = (
            f"{target_year} company master data is not available; use {used_year} record."
        )
    return Company(
        company_code=company_code,
        company_name=_clean(row.get("ShortName")) or company_code,
        company_full_name=_clean(row.get("FullName")),
        exchange=exchange,
        industry_code=_clean(row.get("IndustryCodeD")) or _clean(row.get("IndustryCode")),
        industry_name=_clean(row.get("IndustryNameD")) or _clean(row.get("IndustryName")),
        listing_date=_date(row.get("LISTINGDATE")),
        listing_status=_clean(row.get("LISTINGSTATE")),
        registered_address=_clean(row.get("RegisterAddress")),
        office_address=_clean(row.get("OfficeAddress")),
        secretary=_clean(row.get("Secretary")),
        secretary_tel=_clean(row.get("SecretaryTel")),
        secretary_email=_clean(row.get("SecretaryEmail")),
        website=_clean(row.get("Website")),
        main_business=_clean(row.get("MAINBUSSINESS")),
        company_info_year_requested=target_year,
        company_info_year_used=used_year,
        company_info_fallback=fallback,
        fallback_reason=fallback_reason,
    )


def clear_company_cache() -> None:
    _rows_for_company.cache_clear()
